import openpyxl
import statistics
import random
import os
import json
import requests
import api.service.review_service as review_service
import api.exceptions as api_exceptions

#---------------------------------------------------------------------------
#   Constants
#---------------------------------------------------------------------------
FEATURE_MODELS = ["t-frex-bert-base-uncased", "t-frex-bert-large-uncased", "t-frex-roberta-base", "t-frex-roberta-large", "t-frex-xlnet-base-cased", "t-frex-xlnet-large-cased"]
SENTIMENT_MODELS = ["GPT-3.5"]

# FEATURE_MODELS = ["transfeatex", "t-frex-bert-base-uncased", "t-frex-bert-large-uncased", "t-frex-roberta-base", "t-frex-roberta-large", "t-frex-xlnet-base-cased", "t-frex-xlnet-large-cased"]
# SENTIMENT_MODELS = ["BERT", "BETO","GPT-3.5"]
#---------------------------------------------------------------------------
#---------------------------------------------------------------------------

def set_columns_width(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

def set_benchmark_results(model_dict, results):
        model_dict['benchmark_results']['total_execution_times'].append(results['total_analysis_time'])
        for rev in results['reviews']:
            review_result_dict = {
                'total_review_analysis_time': rev['review_total_analysis_time'],
                'sentences': []
            }
            for sentence in rev['sentences']:
                sentence_dict = {
                        "total_sentence_analysis_times": [],
                        "total_task_analysis_times": []
                }
                if sentence.get('sentence_sentiment_analysis_time', None) is not None:
                    sentence_dict['total_task_analysis_times'].append(sentence.get('sentence_sentiment_analysis_time'))
                if sentence.get('sentence_feature_analysis_time', None) is not None:
                    sentence_dict['total_task_analysis_times'].append(sentence.get('sentence_feature_analysis_time'))
                sentence_dict['total_sentence_analysis_times'].append(sentence.get('sentence_total_analysis_time'))
                review_result_dict['sentences'].append(sentence_dict)
            model_dict['benchmark_results']['reviews'].append(review_result_dict)


def generate_random_reviews(total_reviews=100):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    dataset_path = os.path.join(current_directory, 'big_dataset.json')
    with open(dataset_path, 'r') as dataset:
        data = json.load(dataset)

    all_reviews = []
    for app in data:
        if app.get('reviews', None) is not None:
            for review in app['reviews']:
                all_reviews.append(review)

    random.shuffle(all_reviews)

    selected_reviews = all_reviews[:total_reviews]
    # simplified_reviews = []
    complete_reviews = []

    for rev in selected_reviews:
        # simplified_review = {'reviewId': rev['reviewId'], 'review': rev['review']}
        # simplified_reviews.append(simplified_review)
        complete_reviews.append(rev)
    randomized_dataset = []

    for review_json in complete_reviews:
        randomized_dataset.append(review_service.extract_review_dto_from_json(review_json))    
    for review_dto in randomized_dataset:
        review_service.check_review_splitting(review_dto)

    return randomized_dataset

def send_to_hub_for_performance(reviews, feature_model, sentiment_model, hub_version):
    endpoint_url = os.environ.get('HUB_URL', 'http://127.0.0.1:3002') + '/analyze/performance'
    # api_logger.info(f"[{datetime.now()}]: HUB URL {endpoint_url}")
    
    version = ""
    if hub_version == None: 
        version = 'v0'
    else:
        version = hub_version

    if sentiment_model and feature_model:
        endpoint_url += f'?sentiment_model={sentiment_model}&feature_model={feature_model}&hub_version={version}'
    elif sentiment_model:
        endpoint_url += f'?sentiment_model={sentiment_model}&hub_version={version}'
    elif feature_model:
        endpoint_url += f'?feature_model={feature_model}&hub_version={version}'

    reviews_dict = [review.to_dict() for review in reviews]
    response = requests.post(endpoint_url, json=reviews_dict)

    if response.status_code == 200:
        return json.loads(response.content)
    else:
        # api_logger.info(f"[{datetime.now()}]: HUB unnexpected response {response.status_code} {response}")
        raise api_exceptions.HUBException()

def test_sentiment_models_performance(number_of_iterations, dataset):
    sentiment_results = []
    for sentiment_model in SENTIMENT_MODELS:
        model_dict = {
                    "model": sentiment_model,
                      "task": "Sentiment Analysis",
                      "benchmark_results": {
                          "total_execution_times": [],
                          "reviews": [],      
                        }
                    }
        for _ in range(number_of_iterations):
            hub_response = send_to_hub_for_performance(dataset, None, sentiment_model, None)
            set_benchmark_results(model_dict, hub_response)
        sentiment_results.append(model_dict)
    return sentiment_results

def test_feature_models_performance(number_of_iterations, dataset):
    feature_results = []
    for feature_model in FEATURE_MODELS:
        model_dict = {"model": feature_model,
                      "task": "Feature Extraction",
                      "benchmark_results": {
                          "total_execution_times": [],
                          "reviews": [],      
                        }
                    }
        for _ in range(number_of_iterations):
            hub_response = send_to_hub_for_performance(dataset, feature_model, sentiment_model=None, hub_version=None)
            set_benchmark_results(model_dict, hub_response)
        feature_results.append(model_dict)
    return feature_results

def compute_single_process_benchmark_statistics(benchmark_dict):
    average_results = {}
    for benchmark in benchmark_dict:
        avg_total_execution_time = statistics.mean(benchmark['benchmark_results']['total_execution_times'])
        benchmark_reviews = benchmark['benchmark_results']['reviews']
        review_analysis_times = [review['total_review_analysis_time'] for review in benchmark_reviews]
        avg_review_analysis_time = statistics.mean(review_analysis_times)
        benchmark_review_sentences = [review['sentences'] for review in benchmark_reviews]
        sentence_analysis_times = []
        for benchmark_review_sentence in benchmark_review_sentences:
            for review_result in benchmark_review_sentence:
                for result in review_result['total_sentence_analysis_times']:
                    sentence_analysis_times.append(result)
        avg_sentence_analysis_time = statistics.mean(sentence_analysis_times)
        average_results.update({
            benchmark['model']: {
                'task': benchmark['task'],
                'avg_total_execution_time': avg_total_execution_time,
                'avg_sentence_analysis_time': avg_sentence_analysis_time,
                'avg_review_analysis_time': avg_review_analysis_time
            }
        })
    return average_results

def compute_multiple_process_benchmark_statistics(benchmark_dict):
    average_results = []
    for benchmark in benchmark_dict:
        avg_total_execution_time = statistics.mean(benchmark['benchmark_results']['total_execution_times'])
        benchmark_reviews = benchmark['benchmark_results']['reviews']
        review_analysis_times = [review['total_review_analysis_time'] for review in benchmark_reviews]
        avg_review_analysis_time = statistics.mean(review_analysis_times)
        benchmark_review_sentences = [review['sentences'] for review in benchmark_reviews]
        sentence_analysis_times = []
        for benchmark_review_sentence in benchmark_review_sentences:
            for review_result in benchmark_review_sentence:
                for result in review_result['total_sentence_analysis_times']:
                    sentence_analysis_times.append(result)
        avg_sentence_analysis_time = statistics.mean(sentence_analysis_times)
        average_results.append({
                'feature model': benchmark['feature_model'],
                'sentiment model': benchmark['sentiment_model'],
                'avg_total_execution_time': avg_total_execution_time,
                'avg_sentence_analysis_time': avg_sentence_analysis_time,
                'avg_review_analysis_time': avg_review_analysis_time
            })
    return average_results

def compute_average_sentences_per_review(review_dataset):
    num_sentences = 0
    for rev in review_dataset:
        num_sentences += len(rev.sentences)
    return num_sentences/len(review_dataset)

def single_model_benchmark(performance_workbook, number_of_iterations, review_dataset):
    review_qty = len(review_dataset)
    ws_single_model = performance_workbook.create_sheet(title=f"Single model analysis with {review_qty} reviews")
    ws_single_model.append([
                'Language Model',
                'Language Model task',
                f'Average number of sentences per review',
                f'Average sentence execution time for {number_of_iterations} iterations',
                f'Average review execution time for {number_of_iterations} iterations',
                f'Average total execution time for {review_qty} reviews and {number_of_iterations} iterations'
                ])
    set_columns_width(ws_single_model)

    benchmark_dict = []
    benchmark_dict.extend(test_feature_models_performance(number_of_iterations, review_dataset))
    benchmark_dict.extend(test_sentiment_models_performance(number_of_iterations, review_dataset))
    average_results = compute_single_process_benchmark_statistics(benchmark_dict)
    average_sentences = compute_average_sentences_per_review(review_dataset)
    for res in average_results:
        ws_single_model.append([
                                res,
                                average_results[res]['task'],
                                average_sentences,
                                average_results[res]['avg_sentence_analysis_time'], 
                                average_results[res]['avg_review_analysis_time'],   
                                average_results[res]['avg_total_execution_time']])
    return average_results

def multimodel_single_process_benchmark(performance_workbook, number_of_iterations, single_model_results, review_qty):
    ws_multi_model_single_process = performance_workbook.create_sheet(title=f"Multi model Single Process analysis with {review_qty} reviews")
    ws_multi_model_single_process.append(["Feature Model",
                                        "Sentiment Model", 
                                        f"Average sentence feature extraction time for {number_of_iterations} iterations (seconds)", 
                                        f"Average sentence sentiment analysis time for {number_of_iterations} iterations (seconds)",
                                        f"Average total execution time for {review_qty} reviews and {number_of_iterations} iterations (seconds)"],
                                        )

    
    set_columns_width(ws_multi_model_single_process)
    for feature_model in FEATURE_MODELS:
        feature_model_results = single_model_results[feature_model]
        avg_feature_time = feature_model_results.get('avg_sentence_analysis_time', 0)
        avg_feature_total_execution_time = feature_model_results.get('avg_total_execution_time', 0)
        for sentiment_model in SENTIMENT_MODELS:
            sentiment_model_results = single_model_results[sentiment_model]
            
            avg_total_execution_time = avg_feature_total_execution_time + sentiment_model_results.get('avg_total_execution_time', 0)
            avg_sentiment_time = sentiment_model_results.get('avg_sentence_analysis_time', 0)

            ws_multi_model_single_process.append([feature_model,
                        sentiment_model,
                        avg_feature_time, 
                        avg_sentiment_time, 
                        avg_total_execution_time])

def multimodel_multi_process_benchmark(performance_workbook, number_of_iterations, review_dataset):
    review_qty = len(review_dataset)
    ws_multi_model_multi_process = performance_workbook.create_sheet(title=f"Multi model Multiple Process analysis with {review_qty} reviews")
    ws_multi_model_multi_process.append(["Feature Model",
                                        "Sentiment Model", 
                                        f"Average sentence feature extraction time for {number_of_iterations} iterations (seconds)", 
                                        f"Average sentence sentiment analysis time for {number_of_iterations} iterations (seconds)",
                                        f"Average total execution time for {review_qty} reviews and {number_of_iterations} iterations (seconds)"],
                                        )

    set_columns_width(ws_multi_model_multi_process)

    results = []
    for feature_model in FEATURE_MODELS:
        for sentiment_model in SENTIMENT_MODELS:
            for _ in range(number_of_iterations):
                hub_response = send_to_hub_for_performance(review_dataset, feature_model, sentiment_model, hub_version='v1')
                model_dict = {
                    "feature_model": feature_model,
                    "sentiment_model": sentiment_model,
                    "benchmark_results": {
                          "total_execution_times": [],
                          "reviews": [],      
                        }
                    }
                set_benchmark_results(model_dict, hub_response)
            results.append(model_dict)
    
    average_results = compute_multiple_process_benchmark_statistics(results)
    for res in average_results:
        ws_multi_model_multi_process.append([
                res['feature model'],
                res['sentiment model'],
                res['avg_sentence_analysis_time'], 
                res['avg_review_analysis_time'],   
                res['avg_total_execution_time']])


#---------------------------------------------------------------------------
#   Performance main function
#---------------------------------------------------------------------------

def test_performance(number_of_iterations, dataset_size):
    review_dataset = generate_random_reviews(dataset_size)
    performance_workbook = openpyxl.Workbook()
    # ------- Single Model Analysis -------
    results = single_model_benchmark(performance_workbook, number_of_iterations, review_dataset)
    # ------- Multi Model Analysis -------
    multimodel_single_process_benchmark(performance_workbook, number_of_iterations, results, dataset_size)
    # multimodel_multi_process_benchmark(performance_workbook, number_of_iterations, review_ds)

    output_file = f"{dataset_size}_reviews_{number_of_iterations}_iterations_performance_results.xlsx"
    performance_workbook.save(output_file)
